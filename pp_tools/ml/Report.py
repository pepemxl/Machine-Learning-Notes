# -*- coding: utf-8 -*-
"""
Created on Sun Aug 28 14:07:57 2022

@author: pepem

!pip install minepy
!pip install sklego
"""


import numpy as np 
import pandas as pd

import openpyxl
import os
import time
import platform
import base64
from pprint import pprint
import json
from datetime import datetime

import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go

from dataclasses import dataclass
import mapping

from minepy import MINE
import scipy.stats as stats
from scipy.spatial import distance
import minepy

from sklearn.preprocessing import StandardScaler, PowerTransformer, RobustScaler, MinMaxScaler

from sklearn.cluster import AgglomerativeClustering,KMeans, Birch
from sklearn.neighbors import kneighbors_graph
from sklearn.utils import check_array
from sklearn.metrics.pairwise import pairwise_distances_argmin
from sklearn.metrics import accuracy_score

from sklearn.mixture import GaussianMixture,BayesianGaussianMixture
from sklearn.model_selection import train_test_split


@dataclass
class LinkTarget:
    """
    Class to storage temporal data/APIs features
    """
    menu: str
    title: str
    # date: datetime.datetime
    medio: str
    tipo_medio: str
    nota: str
    link: str

def read_xlsx(filename, header=7, lista_datetimes=None):
    df = pd.read_excel(filename, header=header, parse_date=lista_datetimes)
    return df

def read_data_excel(full_filname):
    wb = openpyxl.load_workbook(full_filname, read_only=True)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print("Processing sheet: ",ws.title)
    ws = wb['tab1']
    lista = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        elemento = LinkTarget(
                menu=row[mapping.MENU],
                title=row[mapping.TITLE],
                # date=datetime.today(),
                medio=row[mapping.MEDIUM],
                tipo_medio=row[mapping.MEDIUM_TYPE],
                nota=row[mapping.NOTE],
                link=row[mapping.LINK])
        lista.append(elemento)
        # date=date=datetime.strptime(row[mapping.DATE],"%Y-%m-%d"),
        # date=row[mapping.DATE],
        # date=datetime.strptime(datetime.today(),"%Y-%m-%d"),
    print(lista[0])
    print(ws.cell(row=5,column=6).value)
    print(ws.cell(row=5,column=7).value)
    print(ws.cell(row=5,column=6).hyperlink)
    # c1 = ws.cell(row=5,column=6)
    
    for value in ws.iter_rows(min_row=3, 
                              max_row=31,
                              min_col=1,
                              max_col=5,
                              values_only=True
                              ):
        print(value)
    for row in ws.rows:
        for cell in row:
            print(cell, cell.value)
        print(row)
    print(ws.cell(row=5,column=6).value)
    print(ws.cell(row=5,column=6).hyperlink)
    print(ws.cell(row=19,column=6).value)
    print(ws.cell(row=19,column=6).hyperlink)


class DataStorage:
    """
        
    """
    def __init__(self):
        self.PATH_INPUT = None
        self.PATH_OUTPUT = None
        self.full_filnename = None
        
    def read_configuration(self):
        pass
    
    def create_path_name(self):
        pass
    
    def create_folder(self, name:str):
        pass


class DataStats:
    """
    
    """
    def __init__(self, df:pd.DataFrame = None, fields_ignore:list = None):
        self.fields = None
        self.fields_ignore = fields_ignore
        self.int_fields = None
        self.float_fields = None
        self.numerical_fields = None
        self.categorical_fields = None


class Report:
    
    def __init__(self, filename:str = None):
        self.data_input_filnename = filename
        
        
        
        
        
if __name__== '__main__':
    print('Hola mundo')
    PATH = "D:/CURSOS/Machine-Learning-Notes/data/"
    filename = "data_homes.csv"
    full_filename = os.path.join(PATH)
    df = pd.read_csv()
    #sample_submission = pd.read_csv("D:/DATA/KAGGLE/tabular-playground-series-jul-2022/sample_submission.csv")
    #data = pd.read_csv("D:/DATA/KAGGLE/tabular-playground-series-jul-2022/data.csv")
    #data.head()
    
    # corr = data.corr().round(2)
    # plt.figure(figsize=(20,10))
    # sns.heatmap(corr, vmin=-1, vmax=1, center=0, square=False, annot=True, cmap="PiYG")
    # plt.show()
    
    # fig = plt.figure(figsize=(15,14))
    # for i, f in enumerate(data.columns):
    #     plt.subplot(6, 5, i+1)
    #     sns.histplot(x=data[f])
    #     plt.title(f'feature: {f}')
    
    # fig.suptitle('Feature distributions', size=20)
    # fig.tight_layout()
    # plt.show()
    print(data.columns)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    